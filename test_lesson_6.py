import csv
import os
from io import TextIOWrapper
from os.path import basename
from zipfile import ZipFile

from openpyxl.reader.excel import load_workbook
from pypdf import PdfReader
import pytest

files_dir = os.path.abspath('files')
resources_path = os.path.abspath('resources')
zip_path = os.path.join(resources_path, 'new_archive.zip')


def test_create_archive():
    with ZipFile(zip_path, "w") as zip_file:
        for file in os.listdir(files_dir):
            add_file = os.path.join(files_dir, file)
            zip_file.write(add_file, basename(add_file))
    files = os.listdir(resources_path)
    assert len(files) == 1, 'one archive must be created'
    assert "new_archive.zip" == files[0], 'filename should be new_archive.zip'


def test_pdf():
    with ZipFile(zip_path) as zf:
        pdf_file = zf.extract("docs-pytest-org-en-latest.pdf")
        reader = PdfReader(pdf_file)
        page = reader.pages[0]
        assert 'pytest Documentation' in page.extract_text(), 'information in pdv file not found'
    os.remove(os.path.abspath('docs-pytest-org-en-latest.pdf'))
    zf.close()


def test_xlsx():
    zf = ZipFile(zip_path)
    with zf.open("file_example_XLSX_50.xlsx") as xlsxfile:
        xlsxfile = load_workbook(xlsxfile)
        sheet = xlsxfile.active
        value = sheet.cell(row=2, column=2).value
        assert value == 'Dulce', 'information in xlsx file not found'
        zf.close()


def test_csv():
    zf = ZipFile(zip_path)
    with zf.open("username.csv") as csvfile:
        csvfile = csv.reader(TextIOWrapper(csvfile))
        list_csv = ["".join(r).replace(";", " ", 3) for r in csvfile]
    assert "booker12 9012 Rachel Booker" in list_csv, 'information in csv file not found'
    zf.close()
    os.remove(os.path.join(resources_path, 'new_archive.zip'))


@pytest.fixture()
def clear_dir():
    os.remove(os.path.join(resources_path, 'new_archive.zip'))
